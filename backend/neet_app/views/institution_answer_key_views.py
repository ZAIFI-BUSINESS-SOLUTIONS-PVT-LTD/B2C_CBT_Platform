"""
Institution answer key upload views.
Handles Excel upload for updating correct answers and recalculating is_correct for test sessions.
"""

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.db import transaction, connection
from django.conf import settings
from neet_app.models import Question, TestAnswer, TestSession, PlatformTest
from neet_app.institution_auth import institution_admin_required
import openpyxl
import json
import logging
from datetime import datetime

logger = logging.getLogger(__name__)


class AnswerKeyValidationError(Exception):
    """Custom exception for answer key validation errors"""
    pass


def parse_answer_key_excel(file_obj):
    """
    Parse Excel file with question numbers and answers.
    
    Expected format:
    | question | answer |
    |----------|--------|
    | 1        | A      |
    | 2        | B      |
    | 3        | 45.6   |
    
    Returns:
        list of dicts: [{'question_num': 1, 'answer': 'A'}, ...]
    """
    try:
        workbook = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
        sheet = workbook.active
    except Exception as e:
        raise AnswerKeyValidationError(f"Failed to read Excel file: {str(e)}")
    
    # Parse headers (first row)
    first_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    
    # Find column indices (case-insensitive)
    question_col_idx = None
    answer_col_idx = None
    
    for idx, header in enumerate(first_row):
        if not header:
            continue
        header_lower = str(header).strip().lower()
        
        if header_lower in ['question', 'q', 'question_number', 'question number', 'q_num', 'qno']:
            question_col_idx = idx
        elif header_lower in ['answer', 'correct_answer', 'correct answer', 'ans']:
            answer_col_idx = idx
    
    if question_col_idx is None:
        raise AnswerKeyValidationError(
            "Missing 'question' column. Excel must have headers: 'question' and 'answer'"
        )
    
    if answer_col_idx is None:
        raise AnswerKeyValidationError(
            "Missing 'answer' column. Excel must have headers: 'question' and 'answer'"
        )
    
    # Parse data rows
    answers = []
    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(cell is None or str(cell).strip() == '' for cell in row):
            continue  # Skip empty rows
        
        try:
            question_num = row[question_col_idx]
            answer = row[answer_col_idx]
            
            # Validate question number
            if question_num is None or str(question_num).strip() == '':
                continue  # Skip rows without question number
            
            try:
                question_num = int(question_num)
            except (ValueError, TypeError):
                raise AnswerKeyValidationError(
                    f"Row {row_num}: Invalid question number '{question_num}'. Must be an integer."
                )
            
            # Validate answer (allow empty for unanswered/pending)
            if answer is None:
                answer = ''
            else:
                answer = str(answer).strip()
            
            answers.append({
                'question_num': question_num,
                'answer': answer,
                'row_num': row_num
            })
            
        except IndexError:
            # Row doesn't have enough columns
            continue
    
    if not answers:
        raise AnswerKeyValidationError("No valid answer data found in Excel file")
    
    # Validate sequential numbering (1, 2, 3, ...)
    question_nums = [a['question_num'] for a in answers]
    expected_nums = list(range(1, len(answers) + 1))
    
    if question_nums != expected_nums:
        missing = set(expected_nums) - set(question_nums)
        duplicates = [num for num in question_nums if question_nums.count(num) > 1]
        
        error_msg = "Question numbering issues: "
        if missing:
            error_msg += f"Missing questions: {sorted(missing)}. "
        if duplicates:
            error_msg += f"Duplicate questions: {sorted(set(duplicates))}. "
        error_msg += "Questions must be numbered sequentially from 1."
        
        raise AnswerKeyValidationError(error_msg)
    
    return answers


def update_correct_answers(institution, test_name, answers):
    """
    Update Question.correct_answer for the given test.
    
    Args:
        institution: Institution instance
        test_name: Test name string
        answers: List of answer dicts with 'answer' field
    
    Returns:
        dict with update statistics
    """
    # Get questions for this test ordered by ID
    questions = list(
        Question.objects.filter(
            institution=institution,
            institution_test_name=test_name
        ).order_by('id')
    )
    
    if not questions:
        raise AnswerKeyValidationError(
            f"No questions found for test '{test_name}' in institution '{institution.name}'"
        )
    
    # Validate count match
    if len(questions) != len(answers):
        raise AnswerKeyValidationError(
            f"Question count mismatch: {len(questions)} questions in database "
            f"but {len(answers)} answers provided in Excel"
        )
    
    # Build list of questions needing update
    to_update = []
    backup_data = []
    
    for question, answer_data in zip(questions, answers):
        old_answer = question.correct_answer or ''
        new_answer = answer_data['answer']
        
        # Normalize for comparison
        old_normalized = str(old_answer).strip()
        new_normalized = str(new_answer).strip()
        
        if old_normalized != new_normalized:
            backup_data.append({
                'id': question.id,
                'old_answer': old_answer,
                'new_answer': new_answer
            })
            question._new_correct_answer = new_answer
            to_update.append(question)
    
    # Update in transaction
    with transaction.atomic():
        if to_update:
            for q in to_update:
                q.correct_answer = q._new_correct_answer
            Question.objects.bulk_update(to_update, ['correct_answer'])
    
    return {
        'total_questions': len(questions),
        'updated_count': len(to_update),
        'backup_data': backup_data
    }


def recalculate_is_correct(question_ids, tolerance=0.01):
    """
    Recalculate TestAnswer.is_correct for given question IDs.
    Uses raw SQL for efficiency (similar to update_is_correct.py).
    
    Args:
        question_ids: List of question IDs to recalculate
        tolerance: Numeric tolerance for NVT comparison
    
    Returns:
        dict with recalculation statistics
    """
    if not question_ids:
        return {'rows_affected': 0}
    
    # Build SQL CASE expression for is_correct calculation
    placeholders = ','.join(['%s'] * len(question_ids))
    
    case_sql = f"""
CASE
  WHEN ta.selected_answer IS NOT NULL AND trim(ta.selected_answer) <> '' THEN
    CASE WHEN upper(trim(ta.selected_answer)) = upper(trim(q.correct_answer::text)) THEN true ELSE false END

  WHEN ta.text_answer IS NOT NULL AND trim(ta.text_answer) <> '' THEN
    CASE
      WHEN q.correct_answer IS NULL THEN false
      WHEN trim(ta.text_answer) ~ '^-?[0-9]+(\.[0-9]+)?$' AND trim(q.correct_answer::text) ~ '^-?[0-9]+(\.[0-9]+)?$' THEN
        CASE WHEN abs( (ta.text_answer::numeric) - (q.correct_answer::numeric) ) <= %s THEN true ELSE false END
      WHEN lower(trim(ta.text_answer)) = lower(trim(q.correct_answer::text)) THEN true
      ELSE false
    END
  ELSE false
END
"""
    
    update_sql = f"""
UPDATE test_answers ta
SET is_correct = {case_sql}
FROM questions q
WHERE ta.question_id = q.id
  AND ta.question_id IN ({placeholders})
"""
    
    with transaction.atomic():
        with connection.cursor() as cursor:
            # Execute update with tolerance first, then question_ids
            cursor.execute(update_sql, [tolerance] + list(question_ids))
            rows_affected = cursor.rowcount
    
    return {'rows_affected': rows_affected}


def update_test_session_statistics(question_ids):
    """
    Recalculate session statistics for sessions containing updated questions.
    
    Args:
        question_ids: List of question IDs
    
    Returns:
        dict with sessions updated count
    """
    if not question_ids:
        return {'sessions_updated': 0}
    
    # Find affected sessions
    affected_sessions = TestSession.objects.filter(
        testanswer__question_id__in=question_ids
    ).distinct()
    
    sessions_updated = 0
    
    for session in affected_sessions:
        # Recalculate statistics from TestAnswer
        answers = TestAnswer.objects.filter(session=session)
        
        correct = answers.filter(is_correct=True).count()
        incorrect = answers.filter(is_correct=False, selected_answer__isnull=False).count()
        # For NVT: if text_answer exists and is_correct=False, count as incorrect
        incorrect += answers.filter(is_correct=False, text_answer__isnull=False).exclude(text_answer='').count()
        
        # Unanswered: neither selected_answer nor text_answer present (or is_correct is None)
        unanswered = answers.filter(
            is_correct__isnull=True
        ).count()
        unanswered += answers.filter(
            selected_answer__isnull=True,
            text_answer__isnull=True
        ).count()
        unanswered += answers.filter(
            selected_answer='',
            text_answer=''
        ).count()
        
        # Update session
        session.correct_answers = correct
        session.incorrect_answers = incorrect
        session.unanswered = unanswered
        session.save(update_fields=['correct_answers', 'incorrect_answers', 'unanswered'])
        
        sessions_updated += 1
    
    return {'sessions_updated': sessions_updated}


@csrf_exempt
@institution_admin_required
@require_http_methods(["POST"])
def upload_answer_key(request):
    """
    Upload answer key Excel and update correct answers + recalculate is_correct.
    
    POST /api/institution-admin/upload-answer-key
    Content-Type: multipart/form-data
    Fields:
        - file: Excel file (.xlsx) with columns: question, answer
        - test_name: Name of the test
    
    Returns: {
        "success": true,
        "test_name": "...",
        "total_questions": 50,
        "updated_answers": 25,
        "recalculated_test_answers": 150,
        "updated_sessions": 5,
        "backup_data": [...]
    }
    """
    try:
        institution = request.institution
        admin = request.institution_admin
        
        # Get form data
        test_name = request.POST.get('test_name', '').strip()
        
        # Get uploaded file
        if 'file' not in request.FILES:
            return JsonResponse({
                'error': 'MISSING_FILE',
                'message': 'No file uploaded'
            }, status=400)
        
        file_obj = request.FILES['file']
        
        # Validate inputs
        if not test_name:
            return JsonResponse({
                'error': 'INVALID_INPUT',
                'message': 'Test name is required'
            }, status=400)
        
        # Validate file extension
        if not file_obj.name.endswith('.xlsx'):
            return JsonResponse({
                'error': 'INVALID_FILE_TYPE',
                'message': 'Only .xlsx files are supported'
            }, status=400)
        
        # Validate file size (10MB limit)
        max_size = 10 * 1024 * 1024
        file_obj.seek(0, 2)
        size = file_obj.tell()
        file_obj.seek(0)
        
        if size > max_size:
            return JsonResponse({
                'error': 'FILE_TOO_LARGE',
                'message': 'File size must be less than 10MB'
            }, status=400)
        
        # Parse answer key Excel
        try:
            answers = parse_answer_key_excel(file_obj)
        except AnswerKeyValidationError as e:
            return JsonResponse({
                'error': 'VALIDATION_ERROR',
                'message': str(e)
            }, status=400)
        
        # Update correct answers
        try:
            update_result = update_correct_answers(institution, test_name, answers)
        except AnswerKeyValidationError as e:
            return JsonResponse({
                'error': 'VALIDATION_ERROR',
                'message': str(e)
            }, status=400)
        
        # Get question IDs for this test
        question_ids = list(
            Question.objects.filter(
                institution=institution,
                institution_test_name=test_name
            ).values_list('id', flat=True)
        )
        
        # Recalculate is_correct for all test answers with these questions
        tolerance = settings.NEET_SETTINGS.get('NVT_NUMERIC_TOLERANCE', 0.01)
        recalc_result = recalculate_is_correct(question_ids, tolerance)
        
        # Update session statistics
        session_result = update_test_session_statistics(question_ids)
        
        logger.info(
            f"Institution {institution.name} (admin: {admin.username}) "
            f"uploaded answer key for test '{test_name}': "
            f"{update_result['updated_count']}/{update_result['total_questions']} answers updated, "
            f"{recalc_result['rows_affected']} test answers recalculated, "
            f"{session_result['sessions_updated']} sessions updated"
        )
        
        return JsonResponse({
            'success': True,
            'test_name': test_name,
            'total_questions': update_result['total_questions'],
            'updated_answers': update_result['updated_count'],
            'recalculated_test_answers': recalc_result['rows_affected'],
            'updated_sessions': session_result['sessions_updated'],
            'backup_data': update_result['backup_data']
        }, status=200)
        
    except Exception as e:
        logger.exception("Error in upload_answer_key")
        return JsonResponse({
            'error': 'SERVER_ERROR',
            'message': 'An unexpected error occurred during answer key upload'
        }, status=500)
