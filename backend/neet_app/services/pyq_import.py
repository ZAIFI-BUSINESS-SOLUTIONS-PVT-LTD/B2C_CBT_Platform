"""
PYQ (Previous Year Question Paper) import service.
Handles Excel file parsing, validation, and creation of PYQ-scoped questions.
Reuses parsing logic from institution_upload.py with PYQ-specific adaptations.
"""

import openpyxl
from typing import Dict, List, Any, Tuple
from django.db import transaction
from django.core.exceptions import ValidationError
from neet_app.models import Institution, Question, Topic, PreviousYearQuestionPaper
import logging

# Import reusable utilities from institution_upload
from .institution_upload import (
    validate_file_size,
    parse_excel_headers,
    parse_excel_rows,
    UploadValidationError
)

logger = logging.getLogger(__name__)


@transaction.atomic
def create_pyq_and_questions(
    institution: Institution,
    pyq_name: str,
    exam_type: str,
    questions_data: List[Dict[str, Any]],
    uploaded_by,
    source_filename: str,
    notes: str = None
) -> Tuple[PreviousYearQuestionPaper, List[Question]]:
    """
    Create a PreviousYearQuestionPaper and associated Question records.
    This is wrapped in a transaction to ensure atomicity.
    
    Args:
        institution: Institution instance
        pyq_name: Name of the PYQ
        exam_type: Exam type (e.g., 'neet', 'jee')
        questions_data: List of validated question dictionaries
        uploaded_by: InstitutionAdmin who uploaded this
        source_filename: Original filename
        notes: Optional notes/description
    
    Returns:
        Tuple of (PreviousYearQuestionPaper, List[Question])
    """
    from neet_app.views.utils import clean_mathematical_text
    
    # Create questions first
    created_questions = []
    
    for q_data in questions_data:
        # Clean mathematical/LaTeX-like content before saving
        try:
            cleaned_question = clean_mathematical_text(q_data.get('question'))
            cleaned_option_a = clean_mathematical_text(q_data.get('option_a'))
            cleaned_option_b = clean_mathematical_text(q_data.get('option_b'))
            cleaned_option_c = clean_mathematical_text(q_data.get('option_c'))
            cleaned_option_d = clean_mathematical_text(q_data.get('option_d'))
            cleaned_explanation = clean_mathematical_text(q_data.get('explanation'))
            cleaned_difficulty = clean_mathematical_text(q_data.get('difficulty')) if q_data.get('difficulty') else None
            cleaned_qtype = clean_mathematical_text(q_data.get('question_type')) if q_data.get('question_type') else None
        except Exception:
            # If cleaning fails, fall back to original values but continue
            logger.exception('Error cleaning question text during PYQ upload; saving raw values')
            cleaned_question = q_data.get('question')
            cleaned_option_a = q_data.get('option_a')
            cleaned_option_b = q_data.get('option_b')
            cleaned_option_c = q_data.get('option_c')
            cleaned_option_d = q_data.get('option_d')
            cleaned_explanation = q_data.get('explanation')
            cleaned_difficulty = q_data.get('difficulty')
            cleaned_qtype = q_data.get('question_type')

        question = Question.objects.create(
            topic=q_data['topic'],
            question=cleaned_question,
            option_a=cleaned_option_a,
            option_b=cleaned_option_b,
            option_c=cleaned_option_c,
            option_d=cleaned_option_d,
            correct_answer=q_data['correct_answer'],
            explanation=cleaned_explanation,
            difficulty=cleaned_difficulty,
            question_type=cleaned_qtype,
            # Optional image fields (may be None)
            question_image=q_data.get('question_image'),
            option_a_image=q_data.get('option_a_image'),
            option_b_image=q_data.get('option_b_image'),
            option_c_image=q_data.get('option_c_image'),
            option_d_image=q_data.get('option_d_image'),
            explanation_image=q_data.get('explanation_image'),
            # Link to institution and PYQ name
            institution=institution,
            institution_test_name=pyq_name,  # Use PYQ name as test name for linking
            exam_type=exam_type
        )
        created_questions.append(question)
    
    # Create PreviousYearQuestionPaper record
    pyq = PreviousYearQuestionPaper.objects.create(
        institution=institution,
        name=pyq_name,
        uploaded_by=uploaded_by,
        source_filename=source_filename,
        question_count=len(created_questions),
        exam_type=exam_type,
        is_active=True,
        notes=notes
    )
    
    logger.info(f"Created PYQ: {pyq_name} with {len(created_questions)} questions")
    
    return pyq, created_questions


def process_pyq_upload(
    file_obj,
    institution: Institution,
    pyq_name: str,
    exam_type: str,
    uploaded_by,
    notes: str = None
) -> Dict[str, Any]:
    """
    Main entry point for processing a PYQ Excel upload.
    
    Args:
        file_obj: File-like object (Django UploadedFile)
        institution: Institution instance
        pyq_name: Name for the PYQ
        exam_type: Exam type (e.g., 'neet', 'jee')
        uploaded_by: InstitutionAdmin who uploaded this
        notes: Optional notes/description
    
    Returns:
        Dict with keys: pyq_id, pyq_name, questions_created, topics_used
    
    Raises:
        UploadValidationError: If validation fails
    """
    try:
        # Get source filename
        source_filename = getattr(file_obj, 'name', 'unknown.xlsx')
        
        # Validate file size
        validate_file_size(file_obj)
        
        # Load workbook
        try:
            workbook = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
        except Exception as e:
            raise UploadValidationError(f"Failed to read Excel file: {str(e)}")
        
        # Get first sheet
        if not workbook.sheetnames:
            raise UploadValidationError("Excel file has no sheets")
        
        sheet = workbook[workbook.sheetnames[0]]
        
        # Parse headers (reuse institution upload logic)
        headers = parse_excel_headers(sheet)
        
        # Parse and validate rows (reuse institution upload logic)
        questions_data = parse_excel_rows(sheet, headers, institution, exam_type)
        
        # Create PYQ and questions in database
        pyq, created_questions = create_pyq_and_questions(
            institution=institution,
            pyq_name=pyq_name,
            exam_type=exam_type,
            questions_data=questions_data,
            uploaded_by=uploaded_by,
            source_filename=source_filename,
            notes=notes
        )
        
        # Get unique topics
        topics_used = list(set(q.topic.name for q in created_questions))
        
        return {
            'success': True,
            'pyq_id': pyq.id,
            'pyq_name': pyq.name,
            'questions_created': len(created_questions),
            'topics_used': topics_used,
            'exam_type': exam_type
        }
        
    except UploadValidationError:
        raise
    except Exception as e:
        logger.exception(f"Unexpected error during PYQ upload: {str(e)}")
        raise UploadValidationError(f"Failed to process PYQ upload: {str(e)}")
