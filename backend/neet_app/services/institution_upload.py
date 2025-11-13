"""
Institution question upload service.
Handles Excel file parsing, validation, and creation of institution-scoped questions and tests.
"""

import openpyxl
from typing import Dict, List, Any, Tuple
from django.db import transaction
from django.core.exceptions import ValidationError
from neet_app.models import Institution, Question, Topic, PlatformTest
import logging
import base64
import binascii
# reuse existing cleaning utilities
from neet_app.views.utils import clean_mathematical_text, normalize_subject

logger = logging.getLogger(__name__)

# Expected column names (case-insensitive mapping)
REQUIRED_COLUMNS = {
    'question_text': ['question_text', 'question', 'q', 'question_stem'],
    'option_a': ['option_a', 'a', 'option1'],
    'option_b': ['option_b', 'b', 'option2'],
    'option_c': ['option_c', 'c', 'option3'],
    'option_d': ['option_d', 'd', 'option4'],
    'correct_answer': ['correct_answer', 'answer', 'correct', 'correct_option'],
    'explanation': ['explanation', 'explain', 'solution'],
    'topic_name': ['topic_name', 'topic', 'subject_topic'],  # MOVED FROM OPTIONAL - Required for proper topic classification
    'subject': ['subject', 'subject_name'],  # ADDED - Required for Topic model and analytics
}

OPTIONAL_COLUMNS = {
    'difficulty': ['difficulty', 'level', 'difficulty_level'],
    'question_type': ['question_type', 'type', 'q_type'],
    'chapter': ['chapter', 'chapter_name', 'chapter_number'],  # ADDED - Optional chapter info
    # Optional image columns for question and options (base64 payload or full data-URI)
    'question_image': ['question_image', 'question image', 'questionimage', 'question_img', 'questionimage_base64'],
    'option_a_image': ['option_a_image', 'option a image', 'optionaimage', 'option_a_img', 'option_a_image_base64'],
    'option_b_image': ['option_b_image', 'option b image', 'optionbimage', 'option_b_img', 'option_b_image_base64'],
    'option_c_image': ['option_c_image', 'option c image', 'optioncimage', 'option_c_img', 'option_c_image_base64'],
    'option_d_image': ['option_d_image', 'option d image', 'optiondimage', 'option_d_img', 'option_d_image_base64'],
    'explanation_image': ['explanation_image', 'explanation image', 'explanationimage', 'explanation_img', 'explanation_image_base64'],
}

MAX_ROWS = 5000  # Maximum questions per upload
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB


class UploadValidationError(Exception):
    """Custom exception for upload validation errors"""
    pass


def validate_file_size(file_obj) -> None:
    """Validate uploaded file size"""
    file_obj.seek(0, 2)  # Seek to end
    size = file_obj.tell()
    file_obj.seek(0)  # Reset to beginning
    
    if size > MAX_FILE_SIZE:
        raise UploadValidationError(f"File size exceeds maximum allowed size of {MAX_FILE_SIZE / (1024*1024)}MB")


def normalize_column_name(col_name: str, mapping: Dict[str, List[str]]) -> str:
    """Map user's column name to our internal standard name"""
    col_lower = str(col_name).strip().lower()
    
    for standard_name, variants in mapping.items():
        if col_lower in variants:
            return standard_name
    
    return None


def parse_excel_headers(sheet) -> Dict[str, int]:
    """
    Parse the first row to find column indices.
    Returns a dict mapping standard column names to their column indices (0-based).
    """
    headers = {}
    first_row = sheet[1]
    
    for idx, cell in enumerate(first_row):
        col_value = cell.value
        if not col_value:
            continue
        
        # Try to match required columns
        standard_name = normalize_column_name(col_value, REQUIRED_COLUMNS)
        if standard_name:
            headers[standard_name] = idx
            continue
        
        # Try to match optional columns
        standard_name = normalize_column_name(col_value, OPTIONAL_COLUMNS)
        if standard_name:
            headers[standard_name] = idx
    
    # Validate all required columns are present
    missing = []
    for req_col in REQUIRED_COLUMNS.keys():
        if req_col not in headers:
            missing.append(req_col)
    
    if missing:
        raise UploadValidationError(f"Missing required columns: {', '.join(missing)}")
    
    return headers


def normalize_correct_answer(answer: str) -> str:
    """
    Normalize correct answer from upload.

    Behavior:
    - If the uploader provided an MCQ-style value (A/B/C/D, option_1, 1, etc.) this returns
      the canonical uppercase letter 'A'|'B'|'C'|'D'.
    - If the uploader provided a numeric answer (integer or float) it returns the numeric string
      (trimmed) so NVT numeric answers like '3', '3.14' are preserved.
    - Otherwise, returns the trimmed string value (useful for descriptive/text NVT answers).

    This function is intentionally permissive to support mixed uploads where some questions
    may be MCQ and others NVT.
    """
    if answer is None:
        raise UploadValidationError("Correct answer cannot be empty")

    raw = str(answer).strip()
    if raw == "":
        raise UploadValidationError("Correct answer cannot be empty")

    low = raw.lower()

    # Handle MCQ letter and common variants
    if low in ['a', 'b', 'c', 'd']:
        return low.upper()
    if low in ['option_a', 'option a', '(a)', '(a)', '1', 'first', 'option_1', 'option 1', '1.0']:
        return 'A'
    if low in ['option_b', 'option b', '(b)', '2', 'second', 'option_2', 'option 2', '2.0']:
        return 'B'
    if low in ['option_c', 'option c', '(c)', '3', 'third', 'option_3', 'option 3', '3.0']:
        return 'C'
    if low in ['option_d', 'option d', '(d)', '4', 'fourth', 'option_4', 'option 4', '4.0']:
        return 'D'

    # If looks like a numeric value (integer or float), keep as-is (trimmed)
    import re
    if re.match(r'^-?\d+(\.\d+)?$', raw):
        return raw

    # Otherwise accept the trimmed string (text answer)
    return raw


def get_or_create_topic(
    topic_name: str, 
    subject: str, 
    exam_type: str, 
    institution: Institution,
    chapter: str = None
) -> Topic:
    """
    Get existing topic or create a new one with explicit subject classification.
    
    Args:
        topic_name: Name of the topic/concept
        subject: Subject classification (Physics, Chemistry, Botany, Zoology)
        exam_type: Exam type (e.g., 'neet', 'jee')
        institution: Institution instance
        chapter: Optional chapter information
    
    Returns:
        Topic instance (created or existing)
    
    Raises:
        UploadValidationError: If subject is invalid
    """
    # Normalize subject using canonical mapping
    normalized = normalize_subject(subject)
    if not normalized:
        raise UploadValidationError(f"Invalid subject: '{subject}'. Must be one of: Physics, Chemistry, Botany, Zoology, Math")
    subject = normalized
    
    # Normalize topic name
    topic_name = topic_name.strip()
    
    # Try to find existing topic with same name, subject, and chapter
    # Using get_or_create with name + subject combination
    topic, created = Topic.objects.get_or_create(
        name=topic_name,
        subject=subject,
        chapter=chapter,
        defaults={
            'icon': '📚',  # Default icon
        }
    )
    
    if created:
        logger.info(f"Created new topic: {topic_name} (Subject: {subject}, Chapter: {chapter or 'N/A'})")
    
    return topic


def parse_excel_rows(sheet, headers: Dict[str, int], institution: Institution, exam_type: str) -> List[Dict[str, Any]]:
    """
    Parse all data rows from the Excel sheet.
    Returns a list of validated question dictionaries.
    """
    questions = []
    row_count = 0
    
    # Start from row 2 (skip header row)
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if row_count >= MAX_ROWS:
            raise UploadValidationError(f"Maximum row limit ({MAX_ROWS}) exceeded")
        
        # Skip completely empty rows
        if all(cell is None or str(cell).strip() == '' for cell in row):
            continue
        
        try:
            # Extract required fields
            question_text = row[headers['question_text']]
            option_a = row[headers['option_a']]
            option_b = row[headers['option_b']]
            option_c = row[headers['option_c']]
            option_d = row[headers['option_d']]
            correct_answer = row[headers['correct_answer']]
            explanation = row[headers['explanation']]
            
            # Validate required fields are not empty
            if not question_text or str(question_text).strip() == '':
                raise UploadValidationError(f"Row {row_idx}: Question text is empty")
            
            if not all([option_a, option_b, option_c, option_d]):
                raise UploadValidationError(f"Row {row_idx}: All four options must be provided")
            
            if not explanation or str(explanation).strip() == '':
                raise UploadValidationError(f"Row {row_idx}: Explanation is empty")
            
            # Normalize correct answer
            correct_answer = normalize_correct_answer(correct_answer)
            
            # Extract and validate REQUIRED topic_name
            topic_name = row[headers['topic_name']]
            if not topic_name or str(topic_name).strip() == '':
                raise UploadValidationError(f"Row {row_idx}: Topic name is required and cannot be empty")
            topic_name = str(topic_name).strip()
            
            # Extract and validate REQUIRED subject
            subject = row[headers['subject']]
            if not subject or str(subject).strip() == '':
                raise UploadValidationError(f"Row {row_idx}: Subject is required and cannot be empty")
            
            # Normalize and validate subject using central helper
            normalized_subject = normalize_subject(subject)
            if not normalized_subject:
                raise UploadValidationError(f"Row {row_idx}: Invalid subject '{subject}'. Must be one of: Physics, Chemistry, Botany, Zoology, Math")
            subject = normalized_subject
            
            # Extract optional fields
            chapter = None
            if 'chapter' in headers:
                chapter_val = row[headers['chapter']]
                if chapter_val:
                    chapter = str(chapter_val).strip()
            
            difficulty = None
            if 'difficulty' in headers:
                difficulty = row[headers['difficulty']]
                if difficulty:
                    difficulty = str(difficulty).strip().capitalize()
            
            question_type = None
            if 'question_type' in headers:
                question_type = row[headers['question_type']]

            # Extract optional image/base64 fields (if present). We normalize by stripping a leading data:<mime>;base64, prefix
            def _normalize_base64_field(val, field_name='image'):
                """Normalize Excel cell value into raw base64 payload or None.

                This function accepts either a full data URI (data:...;base64,...) or
                a raw base64 payload. It strips surrounding whitespace/quotes/newlines
                and validates that the remaining payload is valid base64. On invalid
                input it returns None (and logs a warning) so upload can continue.
                """
                if val is None:
                    return None
                try:
                    # Convert to str and strip surrounding whitespace
                    s = str(val).strip()
                    if not s:
                        return None
                    
                    original_length = len(s)
                    logger.info(f'Processing {field_name}: Original length = {original_length} chars')

                    # If it's a full data URI like data:image/png;base64,AAAA..., strip the prefix
                    if s.startswith('data:'):
                        parts = s.split(',', 1)
                        if len(parts) == 2:
                            s = parts[1]
                            logger.info(f'{field_name}: Stripped data URI prefix, new length = {len(s)}')
                        else:
                            logger.warning(f'{field_name}: Invalid data URI format (no comma found)')
                            return None

                    # Remove common surrounding quotes introduced by Excel or CSV exports
                    if (s.startswith('"') and s.endswith('"')) or (s.startswith("'") and s.endswith("'")):
                        s = s[1:-1]

                    # Remove any whitespace/newlines inside the base64 payload
                    s = ''.join(s.split())

                    if not s:
                        logger.warning(f'{field_name}: Empty after normalization')
                        return None
                    
                    final_length = len(s)
                    logger.info(f'{field_name}: Final base64 length = {final_length} chars, prefix = {s[:20]}...')

                    # Quick validation: try to decode a small prefix safely to ensure valid base64
                    # Do not keep the decoded bytes (to avoid memory pressure); just validate
                    try:
                        # Test with first 512 chars and last 512 chars to ensure both ends are valid
                        base64.b64decode(s[:512], validate=True)
                        if len(s) > 1024:
                            base64.b64decode(s[-512:], validate=True)
                        logger.info(f'{field_name}: Base64 validation passed ✓')
                    except (binascii.Error, ValueError) as e:
                        # Not valid base64 — warn and return None so consumer won't render broken image
                        logger.warning(f'{field_name}: Invalid base64 payload - {str(e)}. First 100 chars: {s[:100]}')
                        return None

                    return s
                except Exception as e:
                    logger.exception(f'Error normalizing base64 {field_name}: {str(e)}')
                    return None

            question_image = None
            if 'question_image' in headers:
                question_image = _normalize_base64_field(row[headers['question_image']], 'question_image')

            option_a_image = None
            if 'option_a_image' in headers:
                option_a_image = _normalize_base64_field(row[headers['option_a_image']], 'option_a_image')

            option_b_image = None
            if 'option_b_image' in headers:
                option_b_image = _normalize_base64_field(row[headers['option_b_image']], 'option_b_image')

            option_c_image = None
            if 'option_c_image' in headers:
                option_c_image = _normalize_base64_field(row[headers['option_c_image']], 'option_c_image')

            option_d_image = None
            if 'option_d_image' in headers:
                option_d_image = _normalize_base64_field(row[headers['option_d_image']], 'option_d_image')

            explanation_image = None
            if 'explanation_image' in headers:
                explanation_image = _normalize_base64_field(row[headers['explanation_image']], 'explanation_image')
            
            # Get or create topic with explicit subject and chapter
            topic = get_or_create_topic(
                topic_name=topic_name,
                subject=subject,
                exam_type=exam_type,
                institution=institution,
                chapter=chapter
            )
            
            # Build question dict
            question_data = {
                'question': str(question_text).strip(),
                'option_a': str(option_a).strip(),
                'option_b': str(option_b).strip(),
                'option_c': str(option_c).strip(),
                'option_d': str(option_d).strip(),
                'correct_answer': correct_answer,
                'explanation': str(explanation).strip(),
                'topic': topic,
                'difficulty': difficulty,
                'question_type': question_type,
                # include any image data (may be None)
                'question_image': question_image,
                'option_a_image': option_a_image,
                'option_b_image': option_b_image,
                'option_c_image': option_c_image,
                'option_d_image': option_d_image,
                'explanation_image': explanation_image,
            }
            
            questions.append(question_data)
            row_count += 1
            
        except IndexError as e:
            raise UploadValidationError(f"Row {row_idx}: Column index error - {str(e)}")
        except Exception as e:
            raise UploadValidationError(f"Row {row_idx}: {str(e)}")
    
    if not questions:
        raise UploadValidationError("No valid questions found in the file")
    
    return questions


@transaction.atomic
def create_institution_test(
    institution: Institution,
    test_name: str,
    exam_type: str,
    questions_data: List[Dict[str, Any]],
    time_limit: int = 180,  # Default 3 hours
    instructions: str = None,
    scheduled_date_time=None
) -> Tuple[PlatformTest, List[Question]]:
    """
    Create a PlatformTest and associated Question records for an institution.
    This is wrapped in a transaction to ensure atomicity.
    
    Returns:
        Tuple of (PlatformTest, List[Question])
    """
    import uuid
    from datetime import datetime
    
    # Generate unique test code
    timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
    unique_suffix = str(uuid.uuid4())[:8]
    test_code = f"INST_{institution.id}_{exam_type.upper()}_{timestamp}_{unique_suffix}"
    
    # Create questions first
    created_questions = []
    topic_ids = set()
    
    for q_data in questions_data:
        # Clean mathematical/LaTeX-like content before saving.
        try:
            cleaned_question = clean_mathematical_text(q_data.get('question'))
            cleaned_option_a = clean_mathematical_text(q_data.get('option_a'))
            cleaned_option_b = clean_mathematical_text(q_data.get('option_b'))
            cleaned_option_c = clean_mathematical_text(q_data.get('option_c'))
            cleaned_option_d = clean_mathematical_text(q_data.get('option_d'))
            cleaned_explanation = clean_mathematical_text(q_data.get('explanation'))
            cleaned_difficulty = clean_mathematical_text(q_data.get('difficulty')) if q_data.get('difficulty') else None
            cleaned_qtype = clean_mathematical_text(q_data.get('question_type')) if q_data.get('question_type') else None
        except Exception:
            # If cleaning fails for some reason, fall back to original values but continue
            logger.exception('Error cleaning question text during institution upload; saving raw values')
            cleaned_question = q_data.get('question')
            cleaned_option_a = q_data.get('option_a')
            cleaned_option_b = q_data.get('option_b')
            cleaned_option_c = q_data.get('option_c')
            cleaned_option_d = q_data.get('option_d')
            cleaned_explanation = q_data.get('explanation')
            cleaned_difficulty = q_data.get('difficulty')
            cleaned_qtype = q_data.get('question_type')

        question = Question.objects.create(
            topic=q_data['topic'],
            question=cleaned_question,
            option_a=cleaned_option_a,
            option_b=cleaned_option_b,
            option_c=cleaned_option_c,
            option_d=cleaned_option_d,
            correct_answer=q_data['correct_answer'],
            explanation=cleaned_explanation,
            difficulty=cleaned_difficulty,
            question_type=cleaned_qtype,
            # Optional image fields (may be None)
            question_image=q_data.get('question_image'),
            option_a_image=q_data.get('option_a_image'),
            option_b_image=q_data.get('option_b_image'),
            option_c_image=q_data.get('option_c_image'),
            option_d_image=q_data.get('option_d_image'),
            explanation_image=q_data.get('explanation_image'),
            # Institution-specific fields
            institution=institution,
            institution_test_name=test_name,
            exam_type=exam_type
        )
        created_questions.append(question)
        topic_ids.add(q_data['topic'].id)
    
    # Create PlatformTest
    platform_test = PlatformTest.objects.create(
        test_name=test_name,
        test_code=test_code,
        test_type='Institution Test',
        description=f"Test created by {institution.name} for {exam_type.upper()}",
        instructions=instructions or f"This is an institution test for {exam_type.upper()}. Answer all questions to the best of your ability.",
        time_limit=time_limit,
        total_questions=len(created_questions),
        selected_topics=list(topic_ids),
        is_active=True,
        is_institution_test=True,
        institution=institution,
        exam_type=exam_type
    )
    # If a scheduled datetime was provided, set it so the test becomes scheduled
    if scheduled_date_time:
        try:
            platform_test.scheduled_date_time = scheduled_date_time
            platform_test.save(update_fields=['scheduled_date_time'])
        except Exception:
            logger.exception('Failed to set scheduled_date_time on PlatformTest')
    
    logger.info(f"Created institution test: {test_code} with {len(created_questions)} questions")
    
    return platform_test, created_questions


def process_upload(
    file_obj,
    institution: Institution,
    test_name: str,
    exam_type: str,
    time_limit: int = 180,
    instructions: str = None,
    scheduled_date_time=None
) -> Dict[str, Any]:
    """
    Main entry point for processing an uploaded Excel file.
    
    Args:
        file_obj: File-like object (Django UploadedFile)
        institution: Institution instance
        test_name: Name for the test
        exam_type: Exam type (e.g., 'neet', 'jee')
        time_limit: Time limit in minutes (default 180)
        instructions: Optional test instructions
    
    Returns:
        Dict with keys: test_id, test_code, questions_created, topics_used
    
    Raises:
        UploadValidationError: If validation fails
    """
    try:
        # Validate file size
        validate_file_size(file_obj)
        
        # Load workbook
        try:
            workbook = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
        except Exception as e:
            raise UploadValidationError(f"Failed to read Excel file: {str(e)}")
        
        # Get first sheet
        if not workbook.sheetnames:
            raise UploadValidationError("Excel file has no sheets")
        
        sheet = workbook[workbook.sheetnames[0]]
        
        # Parse headers
        headers = parse_excel_headers(sheet)
        
        # Parse and validate rows
        questions_data = parse_excel_rows(sheet, headers, institution, exam_type)
        
        # Create test and questions in database
        platform_test, created_questions = create_institution_test(
            institution=institution,
            test_name=test_name,
            exam_type=exam_type,
            questions_data=questions_data,
            time_limit=time_limit,
            instructions=instructions,
            scheduled_date_time=scheduled_date_time
        )
        
        # Get unique topics
        topics_used = list(set(q.topic.name for q in created_questions))
        
        return {
            'success': True,
            'test_id': platform_test.id,
            'test_code': platform_test.test_code,
            'test_name': platform_test.test_name,
            'questions_created': len(created_questions),
            'topics_used': topics_used,
            'exam_type': exam_type,
            'scheduled_date_time': platform_test.scheduled_date_time.isoformat() if platform_test.scheduled_date_time else None
        }
        
    except UploadValidationError:
        raise
    except Exception as e:
        logger.exception("Unexpected error during upload processing")
        raise UploadValidationError(f"Unexpected error: {str(e)}")
