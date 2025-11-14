"""
Offline results upload service.
Handles Excel file parsing, validation, and creation of institution-scoped
questions, tests, student profiles, test sessions, and test answers from
offline (paper-based) test results.
"""

import openpyxl
import logging
import csv
import io
import uuid
from typing import Dict, List, Any, Tuple, Optional
from datetime import datetime, date
from collections import defaultdict
from django.db import transaction
from django.utils import timezone
from django.conf import settings
from neet_app.models import (
    Institution, Question, Topic, PlatformTest, StudentProfile,
    TestSession, TestAnswer
)
from neet_app.services.institution_upload import (
    get_or_create_topic, UploadValidationError,
    normalize_correct_answer as normalize_mcq_answer
)
from neet_app.views.utils import clean_mathematical_text, normalize_subject
from neet_app.utils.student_utils import ensure_unique_student_id

logger = logging.getLogger(__name__)

# Configuration
MAX_ROWS = 5000
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10 MB

# Header mapping (case-insensitive variants)
REQUIRED_COLUMNS = {
    'student_name': ['student_name', 'student name', 'name', 'student', 'full_name', 'fullname'],
    'phone_number': ['phone_number', 'phone', 'mobile', 'mobile_number', 'contact'],
    'test_name': ['test_name', 'test name', 'test', 'testname'],
    'subject': ['subject', 'subject_name', 'subjectname'],
    'topic_name': ['topic_name', 'topic name', 'topic', 'topicname'],
    'question_text': ['question_text', 'question text', 'question', 'q', 'question_stem'],
    'option_a': ['option_a', 'option a', 'a', 'option1', 'option_1'],
    'option_b': ['option_b', 'option b', 'b', 'option2', 'option_2'],
    'option_c': ['option_c', 'option c', 'c', 'option3', 'option_3'],
    'option_d': ['option_d', 'option d', 'd', 'option4', 'option_4'],
    'explanation': ['explanation', 'explain', 'solution', 'answer_explanation'],
    'correct_answer': ['correct_answer', 'correct answer', 'answer', 'correct', 'correct_option'],
    'opted_answer': ['opted_answer', 'opted answer', 'student_answer', 'student answer', 'response', 'student_response'],
}

OPTIONAL_COLUMNS = {
    'email': ['email', 'email_address', 'e-mail', 'student_email'],
    'exam_type': ['exam_type', 'exam type', 'exam', 'examtype'],
    'question_type': ['question_type', 'type', 'q_type', 'questiontype'],
    'time_taken_seconds': ['time_taken_seconds', 'time_taken', 'time', 'duration', 'time_spent'],
    'answered_at': ['answered_at', 'answered at', 'timestamp', 'answer_time'],
    'attempt_status': ['attempt_status', 'status', 'attemptatus'],
    'marks': ['marks', 'score', 'points'],
}


def validate_file_size(file_obj) -> None:
    """Validate uploaded file size"""
    file_obj.seek(0, 2)
    size = file_obj.tell()
    file_obj.seek(0)
    
    if size > MAX_FILE_SIZE:
        raise UploadValidationError(
            f"File size exceeds maximum allowed size of {MAX_FILE_SIZE / (1024*1024):.0f}MB"
        )


def normalize_column_name(col_name: str, mapping: Dict[str, List[str]]) -> Optional[str]:
    """Map user's column name to our internal standard name"""
    if not col_name:
        return None
    col_lower = str(col_name).strip().lower()
    
    for standard_name, variants in mapping.items():
        if col_lower in variants:
            return standard_name
    
    return None


def parse_excel_headers(sheet) -> Dict[str, int]:
    """
    Parse the first row to find column indices.
    Returns a dict mapping standard column names to their column indices (0-based).
    """
    headers = {}
    first_row = sheet[1]
    
    for idx, cell in enumerate(first_row):
        col_value = cell.value
        if not col_value:
            continue
        
        # Try required columns
        standard_name = normalize_column_name(col_value, REQUIRED_COLUMNS)
        if standard_name:
            headers[standard_name] = idx
            continue
        
        # Try optional columns
        standard_name = normalize_column_name(col_value, OPTIONAL_COLUMNS)
        if standard_name:
            headers[standard_name] = idx
    
    # Validate all required columns present
    missing = []
    for req_col in REQUIRED_COLUMNS.keys():
        if req_col not in headers:
            missing.append(req_col)
    
    if missing:
        raise UploadValidationError(f"Missing required columns: {', '.join(missing)}")
    
    return headers


def normalize_phone(phone: str) -> Optional[str]:
    """Normalize phone number (strip spaces, ensure consistent format)"""
    if not phone:
        return None
    # Simple normalization: remove spaces, hyphens, parentheses
    normalized = ''.join(c for c in str(phone).strip() if c.isdigit() or c == '+')
    return normalized if normalized else None


def normalize_answer(answer: str, question_type: Optional[str] = None) -> str:
    """
    Normalize student answer for MCQ or NVT.
    For MCQ: convert to canonical A/B/C/D.
    For NVT: keep as-is (string/number).
    """
    if not answer:
        return None
    
    raw = str(answer).strip()
    if not raw:
        return None
    
    # If NVT, return as-is
    if question_type and question_type.upper() == 'NVT':
        return raw
    
    # Try MCQ normalization
    try:
        return normalize_mcq_answer(raw)
    except:
        # If normalization fails, treat as text answer (NVT)
        return raw


def get_or_create_student(
    student_name: str,
    phone_number: Optional[str],
    email: Optional[str],
    institution: Institution
) -> StudentProfile:
    """
    Find existing student or create a new one.
    Match by phone first, then email, then create new.
    """
    # Try to find by phone
    if phone_number:
        normalized_phone = normalize_phone(phone_number)
        if normalized_phone:
            students = StudentProfile.objects.filter(phone_number=normalized_phone)
            if students.exists():
                return students.first()
    
    # Try to find by email
    if email:
        try:
            return StudentProfile.objects.get(email=email)
        except StudentProfile.DoesNotExist:
            pass
    
    # Create new student
    if not student_name:
        raise UploadValidationError("Student name is required to create new profile")
    
    # Generate placeholder email if missing
    if not email:
        slug = ''.join(c if c.isalnum() else '' for c in student_name.lower())[:20]
        email = f"{slug}@offline.example.com"
    
    # Create minimal profile
    student = StudentProfile(
        full_name=student_name,
        email=email,
        phone_number=normalize_phone(phone_number) if phone_number else None,
        institution=institution,
        is_active=True,
        date_of_birth=date(2000, 1, 1)  # Placeholder DOB for student_id generation
    )
    
    # Generate student_id
    student.student_id = ensure_unique_student_id(student_name, student.date_of_birth)
    
    # Set unusable password (offline students can't login until they register)
    # Set a default password so institution-created offline students can be managed if needed.
    # Requirement: default password = 'inzighted_begins'
    # Keep behaviour minimal: set the password but do not alter other fields.
    student.set_password('inzighted_begins')
    
    student.save()
    logger.info(f"Created new student profile: {student.student_id} - {student_name}")
    
    return student


def parse_and_group_rows(
    sheet,
    headers: Dict[str, int],
    institution: Institution,
    test_name_override: Optional[str] = None
) -> Tuple[Dict[str, List[Dict]], List[Dict], str, str]:
    """
    Parse all data rows and group by student.
    Returns: (student_rows_dict, error_rows, test_name, exam_type)
    """
    student_rows = defaultdict(list)
    error_rows = []
    test_name = test_name_override
    exam_type = None
    row_count = 0
    
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if row_count >= MAX_ROWS:
            error_rows.append({
                'row_number': row_idx,
                'raw_data': str(row)[:200],
                'error_code': 'MAX_ROWS_EXCEEDED',
                'error_message': f'Maximum row limit ({MAX_ROWS}) exceeded'
            })
            break
        
        # Skip empty rows
        if all(cell is None or str(cell).strip() == '' for cell in row):
            continue
        
        try:
            # Extract required fields
            student_name = row[headers['student_name']]
            phone_number = row[headers['phone_number']] if 'phone_number' in headers else None
            test_name_row = row[headers['test_name']]
            subject = row[headers['subject']]
            topic_name = row[headers['topic_name']]
            question_text = row[headers['question_text']]
            option_a = row[headers['option_a']]
            option_b = row[headers['option_b']]
            option_c = row[headers['option_c']]
            option_d = row[headers['option_d']]
            explanation = row[headers['explanation']]
            correct_answer = row[headers['correct_answer']]
            opted_answer = row[headers['opted_answer']]
            
            # Validate required fields
            if not student_name or str(student_name).strip() == '':
                raise ValueError("Student name is required")
            if not question_text or str(question_text).strip() == '':
                raise ValueError("Question text is required")
            if not all([option_a, option_b, option_c, option_d]):
                raise ValueError("All four options are required")
            if not correct_answer:
                raise ValueError("Correct answer is required")
            
            # Extract test_name
            if not test_name:
                if not test_name_row or str(test_name_row).strip() == '':
                    raise ValueError("Test name is required")
                test_name = str(test_name_row).strip()
            
            # Extract optional fields
            email = None
            if 'email' in headers and headers['email'] < len(row):
                email_val = row[headers['email']]
                if email_val:
                    email = str(email_val).strip()
            
            if not exam_type and 'exam_type' in headers and headers['exam_type'] < len(row):
                exam_type_val = row[headers['exam_type']]
                if exam_type_val:
                    exam_type = str(exam_type_val).strip().lower()
            
            question_type = None
            if 'question_type' in headers and headers['question_type'] < len(row):
                q_type_val = row[headers['question_type']]
                if q_type_val:
                    question_type = str(q_type_val).strip()
            
            time_taken = 60  # Default 1 minute
            if 'time_taken_seconds' in headers and headers['time_taken_seconds'] < len(row):
                time_val = row[headers['time_taken_seconds']]
                if time_val:
                    try:
                        time_taken = int(time_val)
                    except:
                        pass
            
            answered_at = None
            if 'answered_at' in headers and headers['answered_at'] < len(row):
                ans_time_val = row[headers['answered_at']]
                if ans_time_val:
                    # Try to parse as datetime
                    if isinstance(ans_time_val, datetime):
                        answered_at = ans_time_val
                    else:
                        try:
                            from django.utils.dateparse import parse_datetime
                            answered_at = parse_datetime(str(ans_time_val))
                        except:
                            pass
            
            # Normalize subject
            normalized_subject = normalize_subject(subject)
            if not normalized_subject:
                raise ValueError(f"Invalid subject: '{subject}'")
            
            # Build row data
            # Clean mathematical expressions for question and options to normalize formatting
            try:
                cleaned_question = clean_mathematical_text(str(question_text).strip())
            except Exception:
                cleaned_question = str(question_text).strip()

            try:
                cleaned_option_a = clean_mathematical_text(str(option_a).strip())
            except Exception:
                cleaned_option_a = str(option_a).strip()

            try:
                cleaned_option_b = clean_mathematical_text(str(option_b).strip())
            except Exception:
                cleaned_option_b = str(option_b).strip()

            try:
                cleaned_option_c = clean_mathematical_text(str(option_c).strip())
            except Exception:
                cleaned_option_c = str(option_c).strip()

            try:
                cleaned_option_d = clean_mathematical_text(str(option_d).strip())
            except Exception:
                cleaned_option_d = str(option_d).strip()

            try:
                cleaned_explanation = clean_mathematical_text(str(explanation).strip()) if explanation else ''
            except Exception:
                cleaned_explanation = str(explanation).strip() if explanation else ''

            row_data = {
                'row_number': row_idx,
                'student_name': str(student_name).strip(),
                'phone_number': normalize_phone(phone_number) if phone_number else None,
                'email': email,
                'subject': normalized_subject,
                'topic_name': str(topic_name).strip(),
                'question_text': cleaned_question,
                'option_a': cleaned_option_a,
                'option_b': cleaned_option_b,
                'option_c': cleaned_option_c,
                'option_d': cleaned_option_d,
                'explanation': cleaned_explanation,
                'correct_answer': normalize_answer(correct_answer, question_type),
                'opted_answer': normalize_answer(opted_answer, question_type),
                'question_type': question_type,
                'time_taken': time_taken,
                'answered_at': answered_at,
            }
            
            # Group by student (use phone as key if available, else name)
            student_key = row_data['phone_number'] if row_data['phone_number'] else row_data['student_name']
            student_rows[student_key].append(row_data)
            row_count += 1
            
        except Exception as e:
            error_rows.append({
                'row_number': row_idx,
                'raw_data': str(row)[:200],
                'error_code': 'PARSE_ERROR',
                'error_message': str(e)
            })
    
    if not test_name:
        raise UploadValidationError("Test name is required (provide via form or in Excel rows)")
    
    # Default exam_type
    if not exam_type:
        exam_type = institution.exam_types[0] if institution.exam_types else 'neet'
    
    return dict(student_rows), error_rows, test_name, exam_type


@transaction.atomic
def create_questions_and_test(
    student_rows: Dict[str, List[Dict]],
    institution: Institution,
    test_name: str,
    exam_type: str
) -> Tuple[PlatformTest, Dict[str, Question], Dict[str, Topic]]:
    """
    Create or get Topics, Questions, and PlatformTest.
    Returns: (platform_test, question_map, topic_map)
    """
    topic_map = {}
    question_map = {}
    
    # Collect unique topics and questions
    unique_topics = set()
    unique_questions = []
    
    for student_key, rows in student_rows.items():
        for row_data in rows:
            topic_key = (row_data['subject'], row_data['topic_name'])
            unique_topics.add(topic_key)
            
            # Build question signature
            q_sig = (
                row_data['subject'],
                row_data['topic_name'],
                row_data['question_text'],
                row_data['option_a'],
                row_data['option_b'],
                row_data['option_c'],
                row_data['option_d'],
            )
            
            # Check if we already have this question
            exists = any(uq[0] == q_sig for uq in unique_questions)
            if not exists:
                unique_questions.append((q_sig, row_data))
    
    # Create topics
    for subject, topic_name in unique_topics:
        topic = get_or_create_topic(
            topic_name=topic_name,
            subject=subject,
            exam_type=exam_type,
            institution=institution,
            chapter=None
        )
        topic_map[(subject, topic_name)] = topic
    
    # Create questions
    for q_sig, row_data in unique_questions:
        topic = topic_map[(row_data['subject'], row_data['topic_name'])]
        
        # Clean text
        try:
            cleaned_question = clean_mathematical_text(row_data['question_text'])
            cleaned_option_a = clean_mathematical_text(row_data['option_a'])
            cleaned_option_b = clean_mathematical_text(row_data['option_b'])
            cleaned_option_c = clean_mathematical_text(row_data['option_c'])
            cleaned_option_d = clean_mathematical_text(row_data['option_d'])
            cleaned_explanation = clean_mathematical_text(row_data['explanation'])
        except:
            cleaned_question = row_data['question_text']
            cleaned_option_a = row_data['option_a']
            cleaned_option_b = row_data['option_b']
            cleaned_option_c = row_data['option_c']
            cleaned_option_d = row_data['option_d']
            cleaned_explanation = row_data['explanation']
        
        # Try to find existing question
        try:
            question = Question.objects.get(
                question=cleaned_question,
                topic=topic,
                option_a=cleaned_option_a,
                option_b=cleaned_option_b,
                option_c=cleaned_option_c,
                option_d=cleaned_option_d,
                institution=institution,
                institution_test_name=test_name
            )
            logger.info(f"Reusing existing question: {question.id}")
        except Question.DoesNotExist:
            # Create new question
            question = Question.objects.create(
                topic=topic,
                question=cleaned_question,
                option_a=cleaned_option_a,
                option_b=cleaned_option_b,
                option_c=cleaned_option_c,
                option_d=cleaned_option_d,
                correct_answer=row_data['correct_answer'],
                explanation=cleaned_explanation,
                question_type=row_data['question_type'],
                institution=institution,
                institution_test_name=test_name,
                exam_type=exam_type
            )
            logger.info(f"Created new question: {question.id}")
        
        question_map[q_sig] = question
    
    # Create PlatformTest
    topic_ids = [t.id for t in topic_map.values()]
    timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
    unique_suffix = str(uuid.uuid4())[:8]
    test_code = f"OFFLINE_{institution.id}_{exam_type.upper()}_{timestamp}_{unique_suffix}"
    
    platform_test = PlatformTest.objects.create(
        test_name=test_name,
        test_code=test_code,
        test_type='Offline Test',
        description=f"Offline test uploaded by {institution.name}",
        instructions=f"Offline test results for {test_name}",
        time_limit=180,  # Default 3 hours
        total_questions=len(unique_questions),
        selected_topics=topic_ids,
        is_active=True,
        is_institution_test=True,
        institution=institution,
        exam_type=exam_type
    )
    
    logger.info(f"Created PlatformTest: {test_code} with {len(unique_questions)} questions")
    
    return platform_test, question_map, topic_map


def evaluate_answer(
    opted_answer: Optional[str],
    correct_answer: str,
    question_type: Optional[str] = None
) -> Optional[bool]:
    """Evaluate if student answer is correct"""
    if not opted_answer:
        return None  # Unanswered
    
    if question_type and question_type.upper() == 'NVT':
        # NVT evaluation
        try:
            # Try numeric comparison
            student_numeric = float(opted_answer)
            correct_numeric = float(correct_answer)
            tolerance = settings.NEET_SETTINGS.get('NVT_NUMERIC_TOLERANCE', 0.01)
            return abs(student_numeric - correct_numeric) <= tolerance
        except (ValueError, TypeError):
            # String comparison
            case_sensitive = settings.NEET_SETTINGS.get('NVT_CASE_SENSITIVE', False)
            if case_sensitive:
                return opted_answer == correct_answer
            else:
                return opted_answer.lower() == correct_answer.lower()
    else:
        # MCQ evaluation
        return str(opted_answer).strip().upper() == str(correct_answer).strip().upper()


@transaction.atomic
def process_student_session(
    student_key: str,
    rows: List[Dict],
    platform_test: PlatformTest,
    question_map: Dict,
    topic_map: Dict,
    institution: Institution
) -> Tuple[Optional[TestSession], List[Dict]]:
    """
    Process one student's rows: create/find student, create session, create answers.
    Returns: (session, errors)
    """
    errors = []
    
    try:
        # Get first row for student info
        first_row = rows[0]
        
        # Get or create student
        student = get_or_create_student(
            student_name=first_row['student_name'],
            phone_number=first_row['phone_number'],
            email=first_row['email'],
            institution=institution
        )
        
        # Determine session timing
        answered_times = [r['answered_at'] for r in rows if r['answered_at']]
        if answered_times:
            start_time = min(answered_times)
            end_time = max(answered_times)
        else:
            start_time = end_time = timezone.now()
        
        # Create TestSession
        session = TestSession.objects.create(
            student_id=student.student_id,
            test_type='platform',
            platform_test=platform_test,
            selected_topics=platform_test.selected_topics,
            time_limit=platform_test.time_limit,
            start_time=start_time,
            end_time=end_time,
            is_completed=True,
            total_questions=len(rows)
        )
        
        # Create TestAnswers
        correct_count = 0
        incorrect_count = 0
        unanswered_count = 0
        total_time = 0
        
        for row_data in rows:
            # Find question
            q_sig = (
                row_data['subject'],
                row_data['topic_name'],
                row_data['question_text'],
                row_data['option_a'],
                row_data['option_b'],
                row_data['option_c'],
                row_data['option_d'],
            )
            
            question = question_map.get(q_sig)
            if not question:
                errors.append({
                    'row_number': row_data['row_number'],
                    'raw_data': str(row_data)[:200],
                    'error_code': 'QUESTION_NOT_FOUND',
                    'error_message': 'Question not found in question map'
                })
                continue
            
            # Evaluate answer
            is_correct = evaluate_answer(
                row_data['opted_answer'],
                row_data['correct_answer'],
                row_data['question_type']
            )
            
            if is_correct is True:
                correct_count += 1
            elif is_correct is False:
                incorrect_count += 1
            else:
                unanswered_count += 1
            
            # Determine answer field
            if row_data['question_type'] and row_data['question_type'].upper() == 'NVT':
                selected_answer = None
                text_answer = row_data['opted_answer']
            else:
                selected_answer = row_data['opted_answer']
                text_answer = None
            
            # Create answer
            TestAnswer.objects.create(
                session=session,
                question=question,
                selected_answer=selected_answer,
                text_answer=text_answer,
                is_correct=is_correct,
                time_taken=row_data['time_taken'],
                answered_at=row_data['answered_at'] or timezone.now()
            )
            
            total_time += row_data['time_taken']
        
        # Update session summary
        session.correct_answers = correct_count
        session.incorrect_answers = incorrect_count
        session.unanswered = unanswered_count
        session.total_time_taken = total_time
        session.save(update_fields=['correct_answers', 'incorrect_answers', 'unanswered', 'total_time_taken'])
        
        # Update subject scores
        try:
            session.update_subject_classification()
            session.calculate_and_update_subject_scores()
        except Exception as e:
            logger.exception(f"Failed to update subject scores for session {session.id}: {e}")
        
        logger.info(f"Created session {session.id} for student {student.student_id}")
        
        return session, errors
        
    except Exception as e:
        logger.exception(f"Failed to process student {student_key}: {e}")
        for row_data in rows:
            errors.append({
                'row_number': row_data['row_number'],
                'raw_data': str(row_data)[:200],
                'error_code': 'STUDENT_PROCESSING_ERROR',
                'error_message': str(e)
            })
        return None, errors


def generate_error_csv(errors: List[Dict]) -> str:
    """Generate CSV content for error report"""
    if not errors:
        return ""
    
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=['row_number', 'raw_data', 'error_code', 'error_message'])
    writer.writeheader()
    writer.writerows(errors)
    
    return output.getvalue()


def process_offline_upload(
    file_obj,
    institution: Institution,
    test_name: Optional[str] = None,
    test_id: Optional[int] = None,
    exam_type: Optional[str] = None
) -> Dict[str, Any]:
    """
    Main entry point for processing offline results upload.
    
    Returns dict with summary and errors.
    """
    try:
        # Validate file size
        validate_file_size(file_obj)
        
        # Load workbook
        try:
            workbook = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
        except Exception as e:
            raise UploadValidationError(f"Failed to read Excel file: {str(e)}")
        
        if not workbook.sheetnames:
            raise UploadValidationError("Excel file has no sheets")
        
        sheet = workbook[workbook.sheetnames[0]]
        
        # Parse headers
        headers = parse_excel_headers(sheet)
        
        # Parse and group rows
        student_rows, parse_errors, final_test_name, final_exam_type = parse_and_group_rows(
            sheet, headers, institution, test_name
        )
        
        if not student_rows:
            raise UploadValidationError("No valid student rows found in file")
        
        # Create questions and test
        platform_test, question_map, topic_map = create_questions_and_test(
            student_rows, institution, final_test_name, final_exam_type or exam_type
        )
        
        # Process each student
        created_sessions = 0
        created_students = 0
        all_errors = parse_errors.copy()
        
        for student_key, rows in student_rows.items():
            session, student_errors = process_student_session(
                student_key, rows, platform_test, question_map, topic_map, institution
            )
            
            if session:
                created_sessions += 1
            
            all_errors.extend(student_errors)
        
        # Generate error CSV if errors exist
        errors_csv = None
        if all_errors:
            errors_csv = generate_error_csv(all_errors)
        
        logger.info(
            f"Offline upload completed: test={platform_test.test_code}, "
            f"sessions={created_sessions}, questions={len(question_map)}, errors={len(all_errors)}"
        )
        
        return {
            'success': True,
            'processed_rows': sum(len(rows) for rows in student_rows.values()),
            'created_sessions': created_sessions,
            'created_students': 0,  # TODO: track new vs existing
            'questions_created': len(question_map),
            'test_id': platform_test.id,
            'test_code': platform_test.test_code,
            'test_name': platform_test.test_name,
            'errors_count': len(all_errors),
            'errors_csv': errors_csv
        }
        
    except UploadValidationError:
        raise
    except Exception as e:
        logger.exception("Unexpected error during offline upload processing")
        raise UploadValidationError(f"Unexpected error: {str(e)}")
