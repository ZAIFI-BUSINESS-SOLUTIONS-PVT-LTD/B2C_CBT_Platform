"""
Unit tests for offline test results upload feature.
Tests the service layer and view endpoint.
"""

import pytest
import io
import openpyxl
from datetime import datetime, date
from django.test import TestCase, RequestFactory
from django.contrib.auth.hashers import make_password
from neet_app.models import (
    Institution, InstitutionAdmin, StudentProfile, 
    Question, Topic, PlatformTest, TestSession, TestAnswer
)
from neet_app.services.offline_results_upload import (
    validate_file_size, parse_excel_headers, normalize_phone,
    normalize_answer, get_or_create_student, process_offline_upload,
    UploadValidationError, MAX_FILE_SIZE
)
from neet_app.views.institution_admin_views import upload_offline_results


class OfflineUploadServiceTests(TestCase):
    """Test cases for offline_results_upload service"""
    
    def setUp(self):
        """Set up test data"""
        # Create institution
        self.institution = Institution.objects.create(
            name="Test Institution",
            code="TEST123",
            exam_types=['neet']
        )
        
        # Create institution admin
        self.admin = InstitutionAdmin.objects.create(
            username="testadmin",
            institution=self.institution,
            is_active=True
        )
        self.admin.set_password("password123")
        self.admin.save()
    
    def test_validate_file_size_valid(self):
        """Test file size validation with valid file"""
        file_obj = io.BytesIO(b"test content")
        try:
            validate_file_size(file_obj)
        except UploadValidationError:
            self.fail("validate_file_size raised UploadValidationError unexpectedly")
    
    def test_validate_file_size_exceeds_limit(self):
        """Test file size validation with oversized file"""
        # Create a file larger than MAX_FILE_SIZE
        large_content = b"x" * (MAX_FILE_SIZE + 1000)
        file_obj = io.BytesIO(large_content)
        
        with self.assertRaises(UploadValidationError) as ctx:
            validate_file_size(file_obj)
        
        self.assertIn("exceeds maximum", str(ctx.exception))
    
    def test_parse_excel_headers_valid(self):
        """Test Excel header parsing with all required columns"""
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Add headers
        headers_row = [
            'student_name', 'phone_number', 'email', 'test_name', 'subject',
            'topic_name', 'question_text', 'option_a', 'option_b', 'option_c',
            'option_d', 'explanation', 'correct_answer', 'opted_answer'
        ]
        ws.append(headers_row)
        
        # Parse headers
        result = parse_excel_headers(ws)
        
        # Verify all required columns found
        self.assertIn('student_name', result)
        self.assertIn('phone_number', result)
        self.assertIn('test_name', result)
        self.assertEqual(result['student_name'], 0)
        self.assertEqual(result['phone_number'], 1)
    
    def test_parse_excel_headers_missing_required(self):
        """Test Excel header parsing with missing required column"""
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Add incomplete headers (missing 'question_text')
        headers_row = [
            'student_name', 'phone_number', 'test_name', 'subject',
            'topic_name', 'option_a', 'option_b', 'option_c', 'option_d'
        ]
        ws.append(headers_row)
        
        with self.assertRaises(UploadValidationError) as ctx:
            parse_excel_headers(ws)
        
        self.assertIn("Missing required columns", str(ctx.exception))
    
    def test_normalize_phone(self):
        """Test phone number normalization"""
        self.assertEqual(normalize_phone("+91 98765 43210"), "+919876543210")
        self.assertEqual(normalize_phone("(123) 456-7890"), "1234567890")
        self.assertEqual(normalize_phone("  9876543210  "), "9876543210")
        self.assertIsNone(normalize_phone(""))
        self.assertIsNone(normalize_phone(None))
    
    def test_normalize_answer_mcq(self):
        """Test MCQ answer normalization"""
        self.assertEqual(normalize_answer("A", "MCQ"), "A")
        self.assertEqual(normalize_answer("option a", None), "A")
        self.assertEqual(normalize_answer("B", None), "B")
    
    def test_normalize_answer_nvt(self):
        """Test NVT answer normalization"""
        result = normalize_answer("42.5", "NVT")
        self.assertEqual(result, "42.5")
        
        result = normalize_answer("glucose", "NVT")
        self.assertEqual(result, "glucose")
    
    def test_get_or_create_student_new(self):
        """Test creating a new student"""
        student = get_or_create_student(
            student_name="Test Student",
            phone_number="9876543210",
            email="test@example.com",
            institution=self.institution
        )
        
        self.assertIsNotNone(student)
        self.assertEqual(student.full_name, "Test Student")
        self.assertEqual(student.phone_number, "9876543210")
        self.assertEqual(student.email, "test@example.com")
        self.assertEqual(student.institution, self.institution)
        self.assertTrue(student.student_id.startswith("STU"))
    
    def test_get_or_create_student_existing_by_phone(self):
        """Test finding existing student by phone"""
        # Create existing student
        existing = StudentProfile.objects.create(
            student_id="STU001",
            full_name="Existing Student",
            email="existing@example.com",
            phone_number="9876543210",
            institution=self.institution,
            date_of_birth=date(2000, 1, 1)
        )
        existing.set_unusable_password()
        existing.save()
        
        # Try to create with same phone
        student = get_or_create_student(
            student_name="Different Name",
            phone_number="9876543210",
            email="different@example.com",
            institution=self.institution
        )
        
        # Should return existing student
        self.assertEqual(student.student_id, "STU001")
        self.assertEqual(student.full_name, "Existing Student")
    
    def test_get_or_create_student_without_email(self):
        """Test creating student without email (generates placeholder)"""
        student = get_or_create_student(
            student_name="No Email Student",
            phone_number="9999999999",
            email=None,
            institution=self.institution
        )
        
        self.assertIsNotNone(student)
        self.assertTrue("@offline.example.com" in student.email)
    
    def _create_sample_excel(self, rows_data):
        """Helper to create Excel file from data"""
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Headers
        headers = [
            'student_name', 'phone_number', 'email', 'test_name', 'exam_type',
            'subject', 'topic_name', 'question_text', 'option_a', 'option_b',
            'option_c', 'option_d', 'explanation', 'correct_answer', 'opted_answer',
            'question_type', 'time_taken_seconds'
        ]
        ws.append(headers)
        
        # Data rows
        for row in rows_data:
            ws.append(row)
        
        # Save to BytesIO
        file_obj = io.BytesIO()
        wb.save(file_obj)
        file_obj.seek(0)
        
        return file_obj
    
    def test_process_offline_upload_basic(self):
        """Test complete upload flow with minimal data"""
        rows_data = [
            [
                'Student One', '9876543210', 'student1@test.com', 'Test 1', 'neet',
                'Physics', 'Mechanics', 'What is velocity?', 'Speed', 'Displacement/Time',
                'Acceleration', 'None', 'Formula v = d/t', 'B', 'B',
                'MCQ', 60
            ],
            [
                'Student One', '9876543210', 'student1@test.com', 'Test 1', 'neet',
                'Chemistry', 'Organic', 'What is benzene?', 'Compound', 'Element',
                'Mixture', 'Ion', 'Benzene is aromatic', 'A', 'A',
                'MCQ', 45
            ],
        ]
        
        file_obj = self._create_sample_excel(rows_data)
        
        result = process_offline_upload(
            file_obj=file_obj,
            institution=self.institution,
            test_name="Test 1",
            exam_type="neet"
        )
        
        # Verify result structure
        self.assertTrue(result['success'])
        self.assertEqual(result['processed_rows'], 2)
        self.assertEqual(result['created_sessions'], 1)
        self.assertEqual(result['questions_created'], 2)
        self.assertIsNotNone(result['test_id'])
        self.assertIsNotNone(result['test_code'])
        
        # Verify database state
        self.assertEqual(PlatformTest.objects.filter(institution=self.institution).count(), 1)
        self.assertEqual(TestSession.objects.count(), 1)
        self.assertEqual(TestAnswer.objects.count(), 2)
        self.assertEqual(Question.objects.filter(institution=self.institution).count(), 2)
        
        # Verify session details
        session = TestSession.objects.first()
        self.assertEqual(session.correct_answers, 2)
        self.assertEqual(session.incorrect_answers, 0)
        self.assertEqual(session.unanswered, 0)
    
    def test_process_offline_upload_multiple_students(self):
        """Test upload with multiple students"""
        rows_data = [
            # Student 1
            [
                'Student One', '9876543210', 'student1@test.com', 'Test 2', 'neet',
                'Physics', 'Mechanics', 'Question 1?', 'A', 'B', 'C', 'D',
                'Explanation 1', 'A', 'A', 'MCQ', 30
            ],
            # Student 2
            [
                'Student Two', '9876543211', 'student2@test.com', 'Test 2', 'neet',
                'Physics', 'Mechanics', 'Question 1?', 'A', 'B', 'C', 'D',
                'Explanation 1', 'A', 'B', 'MCQ', 40
            ],
        ]
        
        file_obj = self._create_sample_excel(rows_data)
        
        result = process_offline_upload(
            file_obj=file_obj,
            institution=self.institution,
            test_name="Test 2",
            exam_type="neet"
        )
        
        # Should have 2 sessions (one per student)
        self.assertEqual(result['created_sessions'], 2)
        self.assertEqual(TestSession.objects.count(), 2)
        
        # Verify correctness
        sessions = TestSession.objects.all()
        correct_session = sessions.get(correct_answers=1)
        incorrect_session = sessions.get(incorrect_answers=1)
        
        self.assertEqual(correct_session.correct_answers, 1)
        self.assertEqual(incorrect_session.correct_answers, 0)
    
    def test_process_offline_upload_missing_test_name(self):
        """Test upload fails when test_name is missing"""
        rows_data = [
            [
                'Student One', '9876543210', 'student1@test.com', '', 'neet',
                'Physics', 'Mechanics', 'Question?', 'A', 'B', 'C', 'D',
                'Explanation', 'A', 'A', 'MCQ', 60
            ],
        ]
        
        file_obj = self._create_sample_excel(rows_data)
        
        with self.assertRaises(UploadValidationError) as ctx:
            process_offline_upload(
                file_obj=file_obj,
                institution=self.institution,
                test_name=None,  # Not provided
                exam_type="neet"
            )
        
        self.assertIn("Test name is required", str(ctx.exception))


class OfflineUploadViewTests(TestCase):
    """Test cases for upload_offline_results view endpoint"""
    
    def setUp(self):
        """Set up test data"""
        self.factory = RequestFactory()
        
        # Create institution
        self.institution = Institution.objects.create(
            name="Test Institution",
            code="TEST123",
            exam_types=['neet']
        )
        
        # Create admin
        self.admin = InstitutionAdmin.objects.create(
            username="testadmin",
            institution=self.institution,
            is_active=True
        )
        self.admin.set_password("password123")
        self.admin.save()
    
    def _create_authenticated_request(self, file_obj, test_name="Test", exam_type="neet"):
        """Helper to create authenticated POST request"""
        from django.core.files.uploadedfile import SimpleUploadedFile
        
        # Create Excel file wrapper
        uploaded_file = SimpleUploadedFile(
            "test.xlsx",
            file_obj.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        # Create request
        request = self.factory.post(
            '/api/institution-admin/upload-results/',
            {
                'file': uploaded_file,
                'test_name': test_name,
                'exam_type': exam_type
            }
        )
        
        # Attach institution and admin (simulating @institution_admin_required decorator)
        request.institution = self.institution
        request.institution_admin = self.admin
        
        return request
    
    def _create_sample_excel(self):
        """Helper to create sample Excel file"""
        wb = openpyxl.Workbook()
        ws = wb.active
        
        headers = [
            'student_name', 'phone_number', 'email', 'test_name', 'exam_type',
            'subject', 'topic_name', 'question_text', 'option_a', 'option_b',
            'option_c', 'option_d', 'explanation', 'correct_answer', 'opted_answer'
        ]
        ws.append(headers)
        
        ws.append([
            'Test Student', '9876543210', 'test@example.com', 'Upload Test', 'neet',
            'Physics', 'Mechanics', 'What is force?', 'Mass x Acceleration', 'Energy',
            'Power', 'Work', 'F = ma', 'A', 'A'
        ])
        
        file_obj = io.BytesIO()
        wb.save(file_obj)
        file_obj.seek(0)
        
        return file_obj
    
    def test_upload_offline_results_success(self):
        """Test successful upload through view endpoint"""
        file_obj = self._create_sample_excel()
        request = self._create_authenticated_request(file_obj, "View Test", "neet")
        
        response = upload_offline_results(request)
        
        self.assertEqual(response.status_code, 201)
        
        # Parse JSON response
        import json
        data = json.loads(response.content)
        
        self.assertTrue(data['success'])
        self.assertGreater(data['processed_rows'], 0)
        self.assertGreater(data['created_sessions'], 0)
        self.assertIn('test_code', data)
    
    def test_upload_offline_results_missing_file(self):
        """Test upload fails when file is missing"""
        request = self.factory.post(
            '/api/institution-admin/upload-results/',
            {'test_name': 'Test', 'exam_type': 'neet'}
        )
        request.institution = self.institution
        request.institution_admin = self.admin
        
        response = upload_offline_results(request)
        
        self.assertEqual(response.status_code, 400)
        
        import json
        data = json.loads(response.content)
        self.assertEqual(data['error'], 'MISSING_FILE')
    
    def test_upload_offline_results_invalid_file_type(self):
        """Test upload fails with non-xlsx file"""
        from django.core.files.uploadedfile import SimpleUploadedFile
        
        txt_file = SimpleUploadedFile(
            "test.txt",
            b"This is not an Excel file",
            content_type="text/plain"
        )
        
        request = self.factory.post(
            '/api/institution-admin/upload-results/',
            {
                'file': txt_file,
                'test_name': 'Test',
                'exam_type': 'neet'
            }
        )
        request.institution = self.institution
        request.institution_admin = self.admin
        
        response = upload_offline_results(request)
        
        self.assertEqual(response.status_code, 400)
        
        import json
        data = json.loads(response.content)
        self.assertEqual(data['error'], 'INVALID_FILE_TYPE')


if __name__ == '__main__':
    pytest.main([__file__, '-v'])
